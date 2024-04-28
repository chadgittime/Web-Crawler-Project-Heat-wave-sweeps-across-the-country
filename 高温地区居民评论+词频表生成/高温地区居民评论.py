import requests  # 导入requests库，用于发送HTTP请求
from fake_useragent import UserAgent  # 导入UserAgent类，用于生成随机User-Agent
import openpyxl  # 导入openpyxl库，用于处理Excel文件

headers = {
    'User-Agent': str(UserAgent().random)  # 使用随机生成的User-Agent作为请求头
}

# 创建一个新的Excel工作簿
wb = openpyxl.Workbook()
ws = wb.active

# 在第一行设置列名
ws['A1'] = '用户名'
ws['B1'] = '评论'
ws['C1'] = 'IP属地'

comment_list = []  # 用于存储评论的列表

num = 0
while True:
    # 发送GET请求获取头条文章的评论数据
    comments = requests.request("GET",
                                "https://www.toutiao.com/article/v2/tab_comments/?aid=24&app_name=toutiao_web&offset=" + str(
                                    num) + "&count=20&group_id=7246280525462274600").json()['data']
    if comments != []:
        for comment in comments:
            user_name = comment['comment']['user_name']  # 获取用户名
            text = comment['comment']['text']  # 获取评论内容
            IP_address = comment['comment']['publish_loc_info']  # 获取IP属地
            comment_list.append([user_name, text, IP_address])  # 将评论信息添加到列表中
        num += 20
    elif comments == []:
        break

# 将评论列表中的数据逐行写入Excel工作簿中
for comment in comment_list:
    ws.append(comment)

wb.save("./高温地区居民评论.xlsx")  # 将工作簿保存为comments.xlsx文件
