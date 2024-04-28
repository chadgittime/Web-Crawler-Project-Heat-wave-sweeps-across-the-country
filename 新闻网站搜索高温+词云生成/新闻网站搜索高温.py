import json
import re

from fake_useragent import UserAgent
import requests
import openpyxl

headers = {
    "Accept": "application/json",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Connection": "keep-alive",
    "Content-Length": "72",
    "Content-Type": "application/json",
    "Cookie": "Hm_lvt_94a1e06bbce219d29285cee2e37d1d26=1686564592,1687009567,1687015896; b-user-id=051ff741-7430-ccf3-6fd3-439b0f6c419f; acw_tc=76b20f7316870176983334584e7fd4de833eb6446e084800486b2fa8fd26bb; ariaDefaultTheme=undefined; Hm_lpvt_94a1e06bbce219d29285cee2e37d1d26=1687018322",
    "Host": "api.thepaper.cn",
    "Origin": "https://www.thepaper.cn",
    "Referer": "https://www.thepaper.cn/",
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-site",
    'User-Agent': str(UserAgent().random),  # 使用随机的用户代理
    "sec-ch-ua": '"Not.A/Brand";v="8", "Chromium";v="114", "Google Chrome";v="114"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"macOS"'
}

wb = openpyxl.Workbook()  # 创建一个新的工作簿
ws = wb.active
ws['A1'] = '新闻标题'  # 在表格的第一行A列添加标题
ws['B1'] = '发布时间'  # 在表格的第一行B列添加标题
news_list = []  # 存储新闻信息的列表

for num in range(1, 101):  # 循环执行100次
    page = {
        'orderType': 3,
        'pageNum': str(num),
        'pageSize': 10,
        'searchType': 1,
        'word': "高温"
    }

    all_news = requests.request("POST", "https://api.thepaper.cn/search/web/news", headers=headers,
                                data=json.dumps(page)).json()['data']['list']
    # 发送 POST 请求获取新闻数据，并将返回的 JSON 数据解析为字典

    for news in all_news:  # 遍历每条新闻
        title = re.sub('<.*?>', '', news['name'])  # 去除标题中的 HTML 标签
        news_list.append([title, news['pubTime']])  # 将标题和发布时间添加到列表中

for news_info in news_list:  # 遍历新闻信息列表
    ws.append(news_info)  # 将新闻信息添加到工作表中

wb.save("./news.xlsx")  # 保存工作簿为 news.xlsx 文件
